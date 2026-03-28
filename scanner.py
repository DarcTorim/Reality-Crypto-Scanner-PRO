import os
import sys
import socket
import ssl
import datetime
import ipaddress
import time
import threading
import http.client
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook, load_workbook
from colorama import init, Fore, Back, Style
from urllib.parse import urlparse

init()

# Глобальные переменные
FILE_LOCK = threading.Lock()
EXCEL_FILE_PATH = None
STOP_FLAG = False

# Порты для проверки
TARGET_PORTS = [
    80, 443,  # Стандарт
    8080, 8443,  # Альтернативы
    2053, 2083, 2087, 2096,  # Панели хостинга
    8880, 8888, 9000, 9443,  # Прокси
    3000, 5000, 8000  # Dev
]

KNOWN_BRANDS = [
    "microsoft.com", "google.com", "apple.com", "cloudflare.com",
    "amazon.com", "facebook.com", "github.com", "akamai.net", "digitalocean.com"
]


def init_excel_file():
    global EXCEL_FILE_PATH
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"Crypto_Domains_{timestamp}.xlsx"
    EXCEL_FILE_PATH = Path.home() / "Desktop" / filename

    wb = Workbook()
    ws = wb.active
    ws.title = "Encrypted Domains"

    # Новые колонки: Encryption Status
    headers = ["Time", "IP", "Subnet", "Open Ports", "Domain", "Is RU?", "Encrypted?", "Source Method"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        cell.alignment = openpyxl.styles.Alignment(horizontal="center", wrap_text=True)

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 40  # Домен
    ws.column_dimensions['F'].width = 8
    ws.column_dimensions['G'].width = 12  # Статус шифрования
    ws.column_dimensions['H'].width = 20

    wb.save(str(EXCEL_FILE_PATH))
    return filename


def check_port(ip, port, timeout=0.8):
    try:
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.settimeout(timeout)
        result = sock.connect_ex((ip, port))
        sock.close()
        return result == 0
    except Exception:
        return False


def get_ssl_info(ip, port, timeout=2):
    """Возвращает список доменов и флаг успеха SSL."""
    domains = []
    context = ssl.create_default_context()
    context.check_hostname = False
    context.verify_mode = ssl.CERT_NONE

    try:
        with socket.create_connection((ip, port), timeout=timeout) as sock:
            with context.wrap_socket(sock, server_hostname=ip) as ssock:
                # Если подключение прошло - значит шифрование есть
                cert = ssock.getpeercert()
                if cert:
                    subject = dict(x[0] for x in cert.get('subject', []))
                    cn = subject.get('commonName')
                    if cn: domains.append(cn)
                    san = cert.get('subjectAltName', [])
                    for type_name, name in san:
                        if type_name == 'DNS':
                            domains.append(name)
        return list(set(domains)), True  # True = Encrypted
    except Exception:
        return [], False


def get_http_info(ip, port, timeout=2):
    """Возвращает домены из редиректов. Шифрование = False."""
    domains = []
    try:
        conn = http.client.HTTPConnection(ip, port, timeout=timeout)
        conn.request("GET", "/", headers={"Host": ip})
        response = conn.getresponse()
        location = response.getheader("Location")
        if location and location.startswith("http"):
            parsed = urlparse(location)
            if parsed.hostname and parsed.hostname != ip:
                domains.append(parsed.hostname)
        conn.close()
        return list(set(domains)), False  # False = Not Encrypted (via this method)
    except Exception:
        return [], False


def get_reverse_dns(ip):
    try:
        return socket.gethostbyaddr(ip)[0]
    except socket.herror:
        return None


def analyze_single_domain(domain, is_encrypted, method):
    """Анализирует один конкретный домен."""
    d_lower = domain.lower()

    # Проверка на .RU
    is_ru = d_lower.endswith('.ru')

    # Проверка на бренды
    is_brand = any(b in d_lower for b in KNOWN_BRANDS)

    # Статус маскировки
    if is_ru and not is_brand:
        mask_status = "✅ Ideal RU"
    elif is_ru and is_brand:
        mask_status = "⚠️ RU Brand"
    elif is_brand:
        mask_status = "⚠️ Global Brand"
    else:
        mask_status = "✅ Unique"

    # Статус шифрования для вывода
    enc_status = "🔒 YES" if is_encrypted else "🔓 NO"

    return is_ru, enc_status, mask_status


def append_to_excel(row_data):
    with FILE_LOCK:
        try:
            wb = load_workbook(str(EXCEL_FILE_PATH))
            ws = wb.active
            ws.append(row_data)
            wb.save(str(EXCEL_FILE_PATH))
        except Exception:
            pass


def scan_single_ip_detailed(ip_str, current_subnet_str):
    if STOP_FLAG:
        return None

    open_ports = []
    found_records = []  # Список словарей: {domain, is_encrypted, method}

    # 1. Сканирование портов
    for port in TARGET_PORTS:
        if check_port(ip_str, port):
            open_ports.append(port)

            # Попытка SSL (Шифрование)
            if port in [443, 8443, 2053, 2083, 2087, 2096, 9443]:
                domains, has_ssl = get_ssl_info(ip_str, port)
                for d in domains:
                    found_records.append({
                        "domain": d,
                        "encrypted": True,
                        "method": f"SSL:{port}"
                    })

            # Попытка HTTP (Без шифрования на этом этапе)
            if port in [80, 8080, 8880, 8888, 3000, 5000, 8000, 9000]:
                domains, _ = get_http_info(ip_str, port)
                for d in domains:
                    # Проверяем, не нашли ли мы этот домен уже через SSL на другом порту
                    existing = [r for r in found_records if r['domain'] == d]
                    if not existing:
                        found_records.append({
                            "domain": d,
                            "encrypted": False,
                            "method": f"HTTP:{port}"
                        })
                    else:
                        # Если уже есть в SSL, обновляем метод, но оставляем шифрование=True
                        pass

                        # 2. PTR Lookup (обычно не шифруется сам по себе, это просто имя)
    ptr = get_reverse_dns(ip_str)
    if ptr:
        # Проверяем, нет ли уже этого домена в списке с шифрованием
        existing = [r for r in found_records if r['domain'] == ptr]
        if not existing:
            found_records.append({
                "domain": ptr,
                "encrypted": False,  # PTR сам по себе не гарантирует HTTPS
                "method": "PTR"
            })

    if not found_records:
        return None

    results_summary = {"ru_count": 0, "encrypted_count": 0, "total": len(found_records)}
    ports_str = ", ".join(map(str, open_ports))
    timestamp = datetime.datetime.now().strftime("%H:%M:%S")

    # Запись КАЖДОГО домена в отдельную строку Excel для детализации
    for rec in found_records:
        is_ru, enc_status, mask_status = analyze_single_domain(rec['domain'], rec['encrypted'], rec['method'])

        if is_ru:
            results_summary["ru_count"] += 1
        if rec['encrypted']:
            results_summary["encrypted_count"] += 1

        ru_col = "🇷🇺 ДА" if is_ru else "Нет"

        row_data = [
            timestamp, ip_str, current_subnet_str, ports_str,
            rec['domain'], ru_col, enc_status, f"{mask_status} ({rec['method']})"
        ]
        append_to_excel(row_data)

        # Вывод в консоль только важных находок
        if rec['encrypted'] and (is_ru or "Unique" in mask_status):
            color = Back.GREEN if is_ru else Back.CYAN
            text = f"🔒 {rec['domain']} ({enc_status}) [{rec['method']}]"
            if is_ru:
                text = f"🇷🇺 {rec['domain']} (RU + Encrypted)"
            print(f"  {color}{Fore.BLACK} {text} {Style.RESET_ALL}")

    return results_summary


def get_next_subnet_base(current_ip_str):
    try:
        ip_obj = ipaddress.ip_address(current_ip_str)
        octets = list(ip_obj.packed)
        octets[2] += 1
        next_ip_int = int.from_bytes(bytes(octets), byteorder='big')
        return str(ipaddress.ip_address(next_ip_int))
    except:
        return None


def print_header():
    os.system('cls' if os.name == 'nt' else 'clear')
    print(Style.BRIGHT + Fore.CYAN + "=" * 70)
    print("   CRYPTO DOMAIN SCANNER")
    print("   Поиск всех доменов с разделением: ЗАШИФРОВАНО / НЕ ЗАШИФРОВАНО")
    print("=" * 70 + Style.RESET_ALL)
    print(Fore.GREEN + "   🔒 - Домен поддерживает SSL (Годен для Reality)")
    print(Fore.RED + "   🔓 - Только HTTP (Не годен для Reality)")
    print(Fore.MAGENTA + "   🇷🇺 - Российский домен (.ru)")
    print(Style.RESET_ALL + "-" * 70)


def main():
    global STOP_FLAG, EXCEL_FILE_PATH

    print_header()

    start_input = input(f"\n{Fore.GREEN}Введите СТАРТОВЫЙ IP:{Style.RESET_ALL} ").strip()

    try:
        ipaddress.ip_address(start_input)
    except ValueError:
        print(Fore.RED + "❌ Неверный формат IP." + Style.RESET_ALL)
        return

    filename = init_excel_file()
    print(f"\n{Fore.GREEN}📄 Файл создан:{Style.RESET_ALL} {filename}")
    print(f"{Fore.CYAN}💡 Совет:{Style.RESET_ALL} В Excel отфильтруйте колонку 'Encrypted?' по значению '🔒 YES'.")
    print(f"{Fore.YELLOW}⛔ Ctrl+C для остановки.{Style.RESET_ALL}\n")

    current_start_ip = start_input
    subnet_count = 0
    total_domains = 0
    total_encrypted = 0
    total_ru = 0

    MAX_WORKERS = 25

    try:
        while True:
            if STOP_FLAG:
                break

            subnet_count += 1
            current_ip_obj = ipaddress.ip_address(current_start_ip)
            base_net = str(current_ip_obj).rsplit('.', 1)[0] + '.0/24'
            network = ipaddress.ip_network(base_net, strict=False)

            print(f"\n{Style.BRIGHT}{Fore.BLUE}[#{subnet_count}] Подсеть: {network}{Style.RESET_ALL}")

            hosts = list(network.hosts())
            active_ips = 0

            with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
                futures = {executor.submit(scan_single_ip_detailed, str(ip), str(network)): ip for ip in hosts}

                for future in as_completed(futures):
                    if STOP_FLAG:
                        break
                    try:
                        res = future.result()
                        if res:
                            active_ips += 1
                            total_domains += res['total']
                            total_encrypted += res['encrypted_count']
                            total_ru += res['ru_count']
                    except Exception:
                        pass

            if active_ips > 0:
                print(f"   {Fore.GREEN}✔ Активных IP: {active_ips} | Найдено доменов: {total_domains}{Style.RESET_ALL}")
            else:
                print(f"   {Fore.WHITE}⚪ Пусто{Style.RESET_ALL}")

            next_ip = get_next_subnet_base(current_start_ip)
            if not next_ip:
                break
            current_start_ip = next_ip
            time.sleep(0.2)

    except KeyboardInterrupt:
        print(f"\n\n{Fore.YELLOW}⛔ Остановка пользователем...{Style.RESET_ALL}")

    finally:
        print(f"\n{Style.BRIGHT}{Fore.CYAN}=== ИТОГИ ==={Style.RESET_ALL}")
        print(f"Всего найдено доменов: {Fore.WHITE}{total_domains}{Style.RESET_ALL}")
        print(f"Из них ЗАШИФРОВАНЫ (SSL): {Fore.GREEN}{total_encrypted} 🔒{Style.RESET_ALL}")
        print(f"Российские (.ru): {Fore.MAGENTA}{total_ru} 🇷🇺{Style.RESET_ALL}")
        if EXCEL_FILE_PATH:
            print(f"Файл: {Fore.GREEN}{EXCEL_FILE_PATH}{Style.RESET_ALL}")
            print("   -> Откройте Excel и отсортируйте колонку 'Encrypted?' чтобы видеть только рабочие для Reality.")
        input("\nНажмите Enter для выхода...")


if __name__ == "__main__":
    import openpyxl.styles

    main()